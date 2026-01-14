"""
EXCEL_WRITER.PY — Injection paramètres dans Excel
==================================================
Flow: model.json + template.xlsx → output.xlsx

Ce qu'il fait:
- Lit model.json (SST)
- Lit template Excel (avec formules)
- Écrit les paramètres dans les cellules mappées
- Sauvegarde le fichier rempli

Ce qu'il ne fait PAS:
- Calculs P&L (Excel fait ses propres calculs)
- Simulation Monte Carlo
- Création de formules
"""

import json
import openpyxl
from pathlib import Path

print("=" * 80)
print("EXCEL_WRITER.PY — Injection paramètres")
print("=" * 80)

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

MODEL_PATH = 'model.json'
TEMPLATE_PATH = '/mnt/user-data/uploads/unit_economics-6.xlsx'
OUTPUT_PATH = 'unit_economics_output.xlsx'

# =============================================================================
# 2. LECTURE MODEL.JSON
# =============================================================================

with open(MODEL_PATH, 'r') as f:
    model = json.load(f)

print(f"\n✓ model.json chargé (version {model['meta']['version']})")

# =============================================================================
# 3. FONCTION POUR TROUVER UNE VALEUR
# =============================================================================

def find_value(key, asset_data):
    """
    Cherche la valeur d'une clé dans config, inputs, ou risks.
    
    Args:
        key: nom du paramètre (ex: 'n_units', 'pct_low')
        asset_data: dict contenant config, inputs, risks
    
    Returns:
        La valeur trouvée, ou None si pas trouvée
    """
    # Chercher dans config
    if key in asset_data['config']:
        return asset_data['config'][key]
    
    # Chercher dans inputs
    if key in asset_data['inputs']:
        return asset_data['inputs'][key]
    
    # Chercher dans risks.revenue
    if key in asset_data['risks']['revenue']:
        return asset_data['risks']['revenue'][key]
    
    # Chercher dans risks.capital
    if key in asset_data['risks']['capital']:
        return asset_data['risks']['capital'][key]
    
    return None

# =============================================================================
# 4. CHARGEMENT TEMPLATE
# =============================================================================

wb = openpyxl.load_workbook(TEMPLATE_PATH)
ws = wb.active

print(f"✓ Template chargé: {TEMPLATE_PATH}")
print(f"  Feuille: {ws.title}")

# =============================================================================
# 5. INJECTION DES PARAMÈTRES
# =============================================================================

print("\nInjection des paramètres...")
print("-" * 60)

total_cells = 0
errors = []

for asset_name in ['immobilier', 'betail', 'embouche']:
    asset_data = model['assets'][asset_name]
    mapping = asset_data['excel_mapping']['cells']
    column = asset_data['excel_mapping']['column']
    
    print(f"\n{asset_name.upper()} (colonne {column}):")
    
    for key, cell in sorted(mapping.items(), key=lambda x: int(x[1][1:])):
        value = find_value(key, asset_data)
        
        if value is None:
            errors.append(f"  ⚠ {cell}: {key} NOT FOUND")
            continue
        
        # Écrire la valeur
        old_value = ws[cell].value
        ws[cell] = value
        total_cells += 1
        
        # Afficher si changement
        if old_value != value:
            print(f"  {cell}: {key} = {value} (était: {old_value})")
        else:
            print(f"  {cell}: {key} = {value}")

# =============================================================================
# 6. AFFICHAGE ERREURS
# =============================================================================

if errors:
    print("\n⚠ ERREURS:")
    for err in errors:
        print(err)

# =============================================================================
# 7. SAUVEGARDE
# =============================================================================

wb.save(OUTPUT_PATH)

print("\n" + "=" * 80)
print(f"✓ {total_cells} cellules écrites")
print(f"✓ Fichier sauvegardé: {OUTPUT_PATH}")
print("=" * 80)

# =============================================================================
# 8. VÉRIFICATION (lecture des formules calculées)
# =============================================================================

print("\nVÉRIFICATION — Formules Excel vs model.json P&L:")
print("-" * 60)

# Recharger pour voir les valeurs
wb2 = openpyxl.load_workbook(OUTPUT_PATH, data_only=False)
ws2 = wb2.active

# Note: data_only=False montre les formules, pas les valeurs calculées
# Pour voir les valeurs calculées, il faut ouvrir dans Excel

print(f"{'Asset':<12} {'Cellule':<10} {'Formule/Valeur':<30}")
print("-" * 60)

checks = [
    ('immobilier', 'C6', 'Capital total'),
    ('immobilier', 'C33', 'Bénéfice/unité'),
    ('betail', 'H6', 'Capital total'),
    ('betail', 'H33', 'Bénéfice/unité'),
    ('embouche', 'M9', 'Capital total'),
    ('embouche', 'M33', 'Bénéfice/veau'),
]

for asset, cell, label in checks:
    val = ws2[cell].value
    print(f"{asset:<12} {cell:<10} {str(val):<30} ({label})")

print("\n💡 Pour vérifier les calculs, ouvrir le fichier dans Excel.")
